#!/usr/bin/env python3
"""
Build a master "financial institutions by state" workbook from FDIC bank data
and NCUA credit union data, with individual branches collapsed down to one
row per (institution, state).

Inputs (all local files you download yourself -- see README.md for where to
get them, since this script does not fetch anything over the network):

  --fdic-institutions   FDIC "Institutions" bulk CSV (one row per bank charter).
                        Must contain a certificate number column (CERT) and a
                        bank name column (NAME).
  --fdic-locations      FDIC "Locations" bulk CSV (one row per branch/office).
                        Must contain a certificate number column (CERT) and a
                        state column (STALP/STATE).
  --ncua-credit-unions  NCUA credit union list/call-report extract. Must
                        contain a credit union name column and a state
                        column. If your NCUA extract is branch-level (one row
                        per branch/office rather than one row per credit
                        union), it will be deduplicated the same way as the
                        FDIC data using the charter/CU-number column when
                        present.

Output:
  --output              Path to the .xlsx workbook to write. Defaults to
                        data/master_institutions.xlsx

The workbook has four sheets:
  1. All Financial Institutions -- State, Institution Name, Type (deduped)
  2. State Summary              -- per-state bank / credit union / total counts
  3. Bank Source Data           -- the deduped bank rows before merging with CUs
  4. Credit Union Source Data   -- the deduped credit union rows

Run `python3 build_master_list.py --help` for all options.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from us_states import US_STATES_AND_DC, abbr_to_name  # noqa: E402


def find_col(df: pd.DataFrame, candidates: list[str]) -> str:
    """Case-insensitive lookup of the first matching column name."""
    lower_map = {c.lower(): c for c in df.columns}
    for candidate in candidates:
        if candidate.lower() in lower_map:
            return lower_map[candidate.lower()]
    raise KeyError(
        f"None of the expected columns {candidates} were found. "
        f"Available columns: {list(df.columns)}"
    )


def load_csv(path: Path) -> pd.DataFrame:
    # dtype=str keeps certificate/charter numbers (e.g. leading zeros) intact.
    return pd.read_csv(path, dtype=str, encoding="latin-1", low_memory=False)


def build_bank_table(institutions_path: Path, locations_path: Path) -> pd.DataFrame:
    institutions = load_csv(institutions_path)
    locations = load_csv(locations_path)

    inst_cert_col = find_col(institutions, ["CERT"])
    inst_name_col = find_col(institutions, ["NAME", "NAMEFULL", "BANKNAME"])

    loc_cert_col = find_col(locations, ["CERT"])
    loc_state_col = find_col(locations, ["STALP", "STATE", "STATEABBR", "STATE_ABBR"])

    active_col = None
    for candidate in ("ACTIVE",):
        if candidate in institutions.columns:
            active_col = candidate
            break
    if active_col:
        institutions = institutions[institutions[active_col].astype(str).isin(["1", "1.0", "True", "true"])]

    inst_lookup = institutions[[inst_cert_col, inst_name_col]].drop_duplicates(subset=[inst_cert_col])
    inst_lookup = inst_lookup.rename(columns={inst_cert_col: "CERT", inst_name_col: "Institution Name"})

    locs = locations[[loc_cert_col, loc_state_col]].rename(columns={loc_cert_col: "CERT", loc_state_col: "StateAbbr"})
    locs = locs.dropna(subset=["CERT", "StateAbbr"]).drop_duplicates(subset=["CERT", "StateAbbr"])

    merged = locs.merge(inst_lookup, on="CERT", how="inner")
    merged["State"] = merged["StateAbbr"].map(abbr_to_name)
    merged["Type"] = "Bank"

    result = merged[["State", "Institution Name", "Type", "CERT"]].drop_duplicates(
        subset=["State", "Institution Name"]
    )
    return result.sort_values(["State", "Institution Name"]).reset_index(drop=True)


def build_credit_union_table(ncua_path: Path) -> pd.DataFrame:
    ncua = load_csv(ncua_path)

    name_col = find_col(ncua, ["CU_NAME", "CREDIT UNION NAME", "CREDIT_UNION_NAME", "NAME", "CUNAME"])
    state_col = find_col(ncua, ["STATE", "PHYSICALADDRESSSTATE", "STALP", "ST"])

    charter_col = None
    for candidate in ("CU_NUMBER", "CHARTER", "CHARTER_NUMBER", "CUNUMBER", "NCUA_CHARTER"):
        if candidate in ncua.columns:
            charter_col = candidate
            break

    cols = [name_col, state_col] + ([charter_col] if charter_col else [])
    subset = ncua[cols].copy()
    rename_map = {name_col: "Institution Name", state_col: "StateAbbr"}
    if charter_col:
        rename_map[charter_col] = "Charter"
    subset = subset.rename(columns=rename_map)
    subset = subset.dropna(subset=["Institution Name", "StateAbbr"])

    dedupe_subset = ["Charter", "StateAbbr"] if charter_col else ["Institution Name", "StateAbbr"]
    subset = subset.drop_duplicates(subset=dedupe_subset)

    # State column may already hold a full state name in some NCUA extracts;
    # only remap through the abbreviation table when it looks like an abbreviation.
    subset["State"] = subset["StateAbbr"].apply(lambda v: abbr_to_name(v) if len(str(v).strip()) <= 2 else str(v).strip())
    subset["Type"] = "Credit Union"

    keep_cols = ["State", "Institution Name", "Type"] + (["Charter"] if charter_col else [])
    result = subset[keep_cols].drop_duplicates(subset=["State", "Institution Name"])
    return result.sort_values(["State", "Institution Name"]).reset_index(drop=True)


def build_state_summary(master: pd.DataFrame) -> pd.DataFrame:
    counts = master.pivot_table(
        index="State", columns="Type", values="Institution Name", aggfunc="count", fill_value=0
    )
    for col in ("Bank", "Credit Union"):
        if col not in counts.columns:
            counts[col] = 0
    counts["Total"] = counts["Bank"] + counts["Credit Union"]
    counts = counts.reindex(US_STATES_AND_DC, fill_value=0)
    counts = counts.rename(columns={"Bank": "Banks", "Credit Union": "Credit Unions"})
    counts = counts[["Banks", "Credit Unions", "Total"]].reset_index().rename(columns={"index": "State"})
    return counts


def autosize_and_bold_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 50)
    ws.freeze_panes = "A2"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--fdic-institutions", type=Path, required=True)
    parser.add_argument("--fdic-locations", type=Path, required=True)
    parser.add_argument("--ncua-credit-unions", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path("data/master_institutions.xlsx"))
    args = parser.parse_args()

    banks = build_bank_table(args.fdic_institutions, args.fdic_locations)
    credit_unions = build_credit_union_table(args.ncua_credit_unions)

    master = pd.concat(
        [
            banks[["State", "Institution Name", "Type"]],
            credit_unions[["State", "Institution Name", "Type"]],
        ],
        ignore_index=True,
    ).sort_values(["State", "Institution Name"]).reset_index(drop=True)

    summary = build_state_summary(master)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="All Financial Institutions", index=False)
        summary.to_excel(writer, sheet_name="State Summary", index=False)
        banks.drop(columns=["CERT"]).to_excel(writer, sheet_name="Bank Source Data", index=False)
        credit_unions.to_excel(writer, sheet_name="Credit Union Source Data", index=False)

        for sheet in writer.sheets.values():
            autosize_and_bold_header(sheet)

    print(f"Wrote {len(master)} institution/state rows ({len(banks)} bank rows, "
          f"{len(credit_unions)} credit union rows) to {args.output}")


if __name__ == "__main__":
    main()
